import random
import tkinter.messagebox

import openpyxl
import random as rnd
import tkinter as tk
import customtkinter as ct
from tkinter import ttk
import glob
import os

Italian_counter = 1
English_counter = 1
new_word_counter=0
selection = 1
rank_counter=0

def check_copies(ws):
    to_copy = []
    to_copy2 = []
    c = ""
    for num in range(1,3):
        for idRow,row in enumerate(
                ws.iter_rows(min_row=2, min_col=num, max_row=ws.max_row, max_col = num, values_only=True)):
            idRow = idRow + 1
            for idCell, cell in enumerate(row):
                if cell:
                    word = (((str(cell).upper()).strip()).replace(" ", "")).split(",")

                    if num == 2:
                        wordB = ((str(ws["A"][idRow].value).upper()).strip()).split(",")
                    else:
                        wordB = ((str(ws["B"][idRow].value).upper()).strip()).split(",")
                    for idRow2, row2 in enumerate(
                            ws.iter_rows(min_row=idRow+2, min_col=num, max_row=ws.max_row, max_col = num,
                                         values_only=True)):
                        idRow2 = idRow2 + idRow + 1
                        print(row, row2)
                        for idCell2, cell2 in enumerate(row2):
                            to_break = 0
                            if cell2:
                                word2 = (((str(cell2).upper()).strip()).replace(" ", "")).split(",")

                                if num == 2:
                                    wordB2= ((str(ws["A"][idRow2].value).upper()).strip()).split(",")
                                else:
                                    wordB2 = ((str(ws["B"][idRow2].value).upper()).strip()).split(",")
                                string = "  Hello,  world!  "

                                for Value_1 in word:
                                    for Value_2 in word2:
                                        Value_1 = Value_1.replace(" ", "")
                                        Value_2 = Value_2.replace(" ", "")
                                        Value_2.strip()
                                        Value_1.strip()
                                        if str(Value_2) == 'Z':
                                            print(Value_1, Value_2)
                                            print(1)
                                        if str(Value_1.strip()) == ' Z':
                                            print(2)
                                        if str(Value_2) == str(Value_1):
                                            print(word, word2)
                                            print(wordB, wordB2)
                                            to_copy = word
                                            to_copy.extend(x for x in word2 if x not in word)
                                            to_copy2 = wordB
                                            to_copy2.extend(x for x in wordB2 if x not in wordB)
                                            #(ws["A"][idRow+2]).value = to_copy
                                            #(ws["B"][idRow+2]).value = to_copy2
                                            c = c[:0]
                                            if num == 1:
                                                print((ws["A"][idRow]).value)
                                                for y in to_copy:
                                                    print(y)
                                                    c = c + str(y) + ','
                                                c = c[:-1]
                                                (ws["A"][idRow]).value = c
                                                print((ws["A"][idRow]).value)
                                                c = c[:0]
                                                for y in to_copy2:
                                                    print(y)
                                                    c = c + str(y) + ','
                                                c = c[:-1]
                                                print("to copy to b" , c)
                                                print((ws["A"][idRow]).value)
                                                (ws["B"][idRow]).value = c
                                                print((ws["A"][idRow]).value)
                                            else:
                                                for y in to_copy:
                                                    print(y)
                                                    c = c + str(y) + ','
                                                c = c[:-1]
                                                (ws["B"][idRow]).value = c
                                                c = c[:0]
                                                for y in to_copy2:
                                                    print(y)
                                                    c = c + str(y) + ','
                                                c = c[:-1]
                                                (ws["A"][idRow]).value = c
                                                print("lol u fucked")
                                            print(idRow, idRow2)
                                            print((ws["A"][idRow]).value)
                                            ws.delete_rows(idRow2 + 1, 1)
                                            print((ws["A"][idRow]).value)
                                            to_break = 1
                                            break

                                    if to_break == 1:
                                        break


def data_check(ws,ws2):

    #to find what column/row is italian words, if its the row need to use the idRow
    asci_chart = 65
    for idRow,row in enumerate(
            ws2.iter_rows(min_row=1, min_col=1, max_row=ws2.max_row, max_col=ws2.max_column, values_only=True)):
        for idCell, cell in enumerate(row):
            if cell == "Italian":
                asci_chart+=idCell
    #copy the needed column

    values_to_check = list(ws2[chr(asci_chart)])

    # to find what column/row is italian words but in the second file
    asci_chart = 65
    for idRow, row in enumerate(
            ws.iter_rows(min_row=1, min_col=1, max_row=ws.max_row, max_col=ws.max_column, values_only=True)):
        for idCell, cell in enumerate(row):
            if cell == "Italian":
                asci_chart += idCell

    # copy the needed column
    index_list=list()
    for idRow2, row2 in enumerate(
            ws2.iter_rows(min_row=1, min_col=1, max_row=ws.max_row, max_col=2, values_only=True)):
        for idRow1, row1 in enumerate(
                ws.iter_rows(min_row=1, min_col=1, max_row=ws.max_row, max_col=2, values_only=True)):
             try:
                 if(row2[0] == row1[0]):
                    index_list.append(idRow2+1)
                    break
             except:
                 continue

    for index, item in enumerate(index_list):
        ws2.delete_rows(item-index, 1)

    for idRow, row in enumerate(
            ws2.iter_rows(min_row=1, min_col=1, max_row=ws2.max_row, max_col=ws2.max_column, values_only=True)):
        ws.append(row)

def show_word(selction, pointer, answer):

    global rank_counter
    global new_word_counter
    global Italian_counter
    global  English_counter
    to_con = 0
    Index = 0
    if Italian_counter >= len(ws["C"]):
        Italian_counter = 1
    if English_counter >= len(ws["C"]):
        English_counter = 1
    if(answer):
        if selction == 0:
            Language_counter = Italian_counter
        else:
            Language_counter = English_counter

        if rank_counter >= 10:
            rank_counter = 0
        #what if there no more less then 3
        print("this is italian", Italian_counter, "this is english", English_counter)
        print("this is rank", rank_counter)
        if rank_counter <= 4:
            for i in range(len(ws["C"])-1-Language_counter):
                if(ws["C"][Language_counter+i].value <= 3):
                    Index = i
                    to_con = 1
                    print(11)
                    break
        if (rank_counter > 4 and rank_counter < 8) or to_con == 0:
            for i in range(len(ws["C"]) - 1 - Language_counter):
                if (ws["C"][Language_counter + i].value <= 7):
                    Index = i
                    to_con = 1
                    print(22)
                    break
        if rank_counter >= 8 or to_con == 0:
            for i in range(len(ws["c"])-1 -Language_counter):
                if(ws["C"][Language_counter + i].value >=0):
                    to_con = 1
                    Index = i
                    break
        print("this is Index" , Index)
        if to_con == 1:
            rank_counter += 1

    #need to check if word exists
    if Italian_counter + Index >= len(ws["C"]):
        Italian_counter = 1
    if English_counter + Index >= len(ws["C"]):
        English_counter = 1
    if selction == pointer:
        if selction == 0:
            word = (ws[chr(selction+65)][Italian_counter+Index].value)
            new_word_counter = Italian_counter + Index
            Italian_counter = Italian_counter + Index
        else:
            word = (ws[chr(selction+65)][English_counter+Index].value)
            new_word_counter = English_counter + Index
            English_counter = English_counter + Index
    else:
        if selction == 0:
            word = (ws[chr(pointer + 65)][Italian_counter+Index].value)
            new_word_counter = Italian_counter + Index
            Italian_counter = Italian_counter + Index
        else:
            word = (ws[chr(pointer + 65)][English_counter+Index].value)
            new_word_counter = English_counter + Index
            English_counter = English_counter + Index
    return (word)

def help(selection, ans_lab):

    word = str(show_word(selection, 1 - selection, 0))
    ans_lab.configure(bg_color='transparent',
                      text = "the correct answer is: " + word,
                      font=("Arial", 25, "bold"))

def answer(Entry_answer, selection, ans_lab, user):

    check = 0
    word = str(show_word(selection, 1 - selection, 0)).split(",")
    for i in word:
        if (i.strip()).upper() ==  ((Entry_answer.get()).strip()).upper():
            ans_lab.configure(text="Correct", bg_color="green")
            check = 1
            if not ws["C"][new_word_counter].value:
                ws["C"][new_word_counter].value = 5
            elif ws["C"][new_word_counter].value == 0:
                ws["C"][new_word_counter].value = 5
                break
            else:
                ws["C"][new_word_counter].value += 1
    if check == 0:
        if not ws["C"][new_word_counter].value:
            ws["C"][new_word_counter].value = 1
        elif ws["C"][new_word_counter].value == 0:
            ws["C"][new_word_counter].value = 1
        else:
            ws["C"][new_word_counter].value -= 1
        ans_lab.configure(text="Incorrect", bg_color="red")

    df.save(user + ".xlsx")

def next_word(selection, lab, ans_lab):

    ans_lab.configure(text="choose one option:", bg_color='transparent')
    global Italian_counter, English_counter

    if selection == 0:
        Italian_counter += 1
    else:
        English_counter+=1

    word = show_word(selection, selection, 1)
    lab.configure(text=word)

def check_user_exists(check):
    folder_path = os.path.dirname(os.path.abspath(__file__))
    return ([os.path.basename(file) for file in glob.glob(os.path.join(folder_path, check + '.xlsx'))])

def new_user(user_entry):

    name_of_sheet = str(user_entry.get())
    if check_user_exists(name_of_sheet):
        tkinter.messagebox.showinfo(message="User with this name already exists")
    else:
        workbook = openpyxl.Workbook()
        # Get the active sheet
        sheet = workbook.active
        # Write data to the sheet
        sheet["A1"] = "Italian"
        sheet["B1"] = "English"
        # Save the workbook
        workbook.save(name_of_sheet + ".xlsx")
        tkinter.messagebox.showinfo(message="successfully added new user")
        Gui(1, name_of_sheet)

def hello_window():
    ct.set_appearance_mode("dark")
    ct.set_default_color_theme("green")
    m = ct.CTk()
    m.geometry("800x300")
    m.title("Pizza pasta")

    frame_log = ct.CTkFrame(m)
    frame_name = ct.CTkFrame(m)
    frame_name.columnconfigure(0, weight=1)
    frame_name.columnconfigure(1, weight=1)
    users = check_user_exists('*')

    for i in range(len(users)):
        users[i] = str(users[i]).replace(".xlsx","")

    #if there are already existing users/xlsx files
    if len(users) > 0:
        ct.CTkLabel(m,
                    text="choose from the avalible users:",
                    font=("Arial", 20, "bold"),
                    ).pack(pady=15)
        for i in range(len(users)):
            frame_log.columnconfigure(i, weight=1)
        for i in range(len(users)):
            ct.CTkButton(frame_log,
                         text= str(users[i]),
                         font=("Arial", 20, "bold"),
                         command=lambda i=i : Gui(random.randint(0,1), users[i]),
                         ).grid(row=0, column=i)

        # if we want to creat a new user
        ct.CTkLabel(m,
                    text="or create a new one",
                    font=("Arial", 20, "bold"),
                    ).place(relx=0.5, rely=0.44, anchor='center')

    #no files so creat a new user
    else:
        frame_log.columnconfigure(0, weight=1)
        frame_log.columnconfigure(1, weight=1)
        ct.CTkLabel(m,
                 text="Ciao\n "
                      "please write your name\n",
                 font=("Arial", 20, "bold"),
                 ).pack(pady=18)

    frame_log.pack(fill='y')

    #Entry and Label for new user
    user_label = ct.CTkLabel(m,text="name:",font=("Arial", 20, "bold"),)
    user_label.place(relx=0.41, rely=0.55, anchor='center')
    user_entry = ct.CTkEntry(m,font=("Arial", 20, "bold"))
    user_entry.place(relx=0.54, rely=0.55, anchor='center')

    ct.CTkButton(m,
                 text="Creat a new user",
                 font=("Arial", 18, "bold"),
                 command = lambda: new_user(user_entry)
                 ).place(relx=0.51, rely=0.68, anchor='center')
    m.mainloop()

def change_Gui(user_selection, lab, Entry_answer, ans_lab):
    global  selection
    selection = user_selection
    word = show_word(selection, selection, 0)
    lab.configure(text=word, font=("Arial", 15, "bold"))
    lab.grid(column = 1, row = 1 - selection)
    Entry_answer.configure(font=("Arial", 15, "bold"))
    Entry_answer.grid(column = 1, row=selection)
    ans_lab.configure(text="choose one option:", bg_color='transparent')

def open_worksheet(user):

    global ws
    global df
    df = openpyxl.load_workbook(user + ".xlsx")
    ws = df.active

def Gui(user_selection, user):
    global English_counter
    global Italian_counter
    global selection
    selection = user_selection

    open_worksheet(user)

    text_user = "ciao, " + user
    ct.set_appearance_mode("dark")
    ct.set_default_color_theme("blue")
    m=ct.CTk()
    m.geometry("900x600")
    m.title("pasta pizza nikol")
    # frames
    ct.CTkLabel(m, height = 100, text=text_user, font = ("Arial", 25, "bold")).pack()



    frame = ct.CTkFrame(m)
    frame.rowconfigure(0, weight=1)
    frame.rowconfigure(1, weight=1)
    frame.columnconfigure(0,weight=1)
    frame.columnconfigure(1,weight=1)
    frame.pack(fill='x')
    # Answer_frame
    A_frame = ct.CTkFrame(m, fg_color=None)
    A_frame.pack(fill='x')
    # B_Frames
    button_frame = ct.CTkFrame(m, fg_color=None)
    button_frame.columnconfigure(0, weight=1)
    button_frame.columnconfigure(1, weight=1)
    button_frame.columnconfigure(2, weight=1)
    button_frame.rowconfigure(0,weight=1)
    button_frame.rowconfigure(1, weight=1)
    button_frame.pack(fill='x')


    # Labels
    ct.CTkLabel(frame, text="English", font = ("Arial", 15, "bold"), height = 100).grid(column=0,row=0)
    ct.CTkLabel(frame, text="Italian", font = ("Arial", 18, "bold"), height = 100).grid(column=0,row=1)
    lab = ct.CTkLabel(frame)
    Entry_answer = ct.CTkEntry(frame)

    # label that depends on the answer
    ans_lab = ct.CTkLabel(A_frame, text="press any button", height=100, font=("Arial", 15, "bold"))
    ans_lab.pack(fill="x")
    Italian_counter = rnd.randint(1, len(ws["A"])-1)
    English_counter = rnd.randint(1, len(ws["A"])-1)
    change_Gui(selection,lab,Entry_answer, ans_lab)

    #upper Buttons
    help_b = ct.CTkButton(
        button_frame,
        text = "Help",
        font=("Arial", 20, "bold"),
        height=100,
        command=lambda: help(selection, ans_lab))
    help_b.grid(column=0,row=0, sticky="WE")

    answer_b = ct.CTkButton(
        button_frame,
        text="Check",
        font=("Arial", 20, "bold"),
        height=100,
        command=lambda: answer(Entry_answer, selection, ans_lab,user))
    answer_b.grid(column=1,row=0, sticky="WE")

    next_b = ct.CTkButton(
        button_frame,
        text="Next word",
        font=("Arial", 20, "bold"),
        height=100,
        command=lambda: next_word(selection, lab, ans_lab))
    next_b.grid(column=2, row=0, sticky="WE")

    #lower button
    ct.CTkButton(button_frame,
                 text="Italian",
                 height=35,
                 font=("Arial", 20, "bold"),
                 fg_color = ("#2CC985", "#2FA572"),
                 command=lambda: change_Gui(0, lab, Entry_answer, ans_lab)
                 ).grid(column=0, row=1, sticky="WE")

    ct.CTkButton(button_frame,
                 text="English",
                 font=("Arial", 20, "bold"),
                 height=35,
                 fg_color=("#2CC985", "#2FA572"),
                 command=lambda: change_Gui(1, lab, Entry_answer, ans_lab)
                 ).grid(column=1, row=1, sticky="WE")

    ct.CTkButton(button_frame,
                 text="Random",
                 height=35,
                 fg_color = ("#2CC985", "#2FA572"),
                 font=("Arial", 20, "bold"),
                 command=lambda: change_Gui(rnd.randint(0, 1), lab, Entry_answer, ans_lab)
                 ).grid(column=2, row=1, sticky="WE")

    df.save(user + ".xlsx")
    m.mainloop()

if __name__ == '__main__':
    global ws
    global df
    hello_window()
    a = input()
    open_worksheet("test1")
    b = input()
    #check_copies(ws)
    df.save("test1.xlsx")
# See PyCharm help at https://www.jetbrains.com/help/pycharm/
